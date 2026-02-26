#!/usr/bin/env python3
"""Genera un file timesheet.xlsx professionale con dashboard analitica."""

from datetime import date, timedelta
from openpyxl import Workbook
from openpyxl.styles import (
    Font, PatternFill, Alignment, Border, Side, numbers
)
from openpyxl.utils import get_column_letter
from openpyxl.chart import BarChart, PieChart, Reference
from openpyxl.chart.series import DataPoint
from openpyxl.chart.label import DataLabelList
from openpyxl.worksheet.datavalidation import DataValidation

# ---------------------------------------------------------------------------
# Costanti stile
# ---------------------------------------------------------------------------
COLOR_HEADER = "2C3E50"
COLOR_GREEN = "27AE60"
COLOR_ORANGE = "E67E22"
COLOR_WEEKEND = "F5F5F5"
COLOR_ALT_ROW = "F8F9FA"
COLOR_WHITE = "FFFFFF"
COLOR_LIGHT_GREEN = "E8F5E9"
COLOR_LIGHT_ORANGE = "FFF3E0"

FONT_TITLE = Font(name="Calibri", size=18, bold=True, color=COLOR_WHITE)
FONT_SUBTITLE = Font(name="Calibri", size=11, color="666666")
FONT_HEADER = Font(name="Calibri", size=10, bold=True, color=COLOR_WHITE)
FONT_TOTAL_LABEL = Font(name="Calibri", size=10, bold=True, color=COLOR_WHITE)
FONT_TOTAL_VALUE = Font(name="Calibri", size=10, bold=True, color=COLOR_WHITE)
FONT_DATA = Font(name="Calibri", size=10)
FONT_KPI_TITLE = Font(name="Calibri", size=9, color="666666")
FONT_KPI_VALUE = Font(name="Calibri", size=16, bold=True)
FONT_DASH_HEADER = Font(name="Calibri", size=10, bold=True, color=COLOR_WHITE)
FONT_DASH_SECTION = Font(name="Calibri", size=13, bold=True, color=COLOR_HEADER)

FILL_HEADER = PatternFill("solid", fgColor=COLOR_HEADER)
FILL_GREEN = PatternFill("solid", fgColor=COLOR_GREEN)
FILL_WEEKEND = PatternFill("solid", fgColor=COLOR_WEEKEND)
FILL_ALT = PatternFill("solid", fgColor=COLOR_ALT_ROW)
FILL_WHITE = PatternFill("solid", fgColor=COLOR_WHITE)
FILL_LIGHT_GREEN = PatternFill("solid", fgColor=COLOR_LIGHT_GREEN)
FILL_LIGHT_ORANGE = PatternFill("solid", fgColor=COLOR_LIGHT_ORANGE)
FILL_ORANGE = PatternFill("solid", fgColor=COLOR_ORANGE)

THIN_BORDER = Border(
    left=Side(style="thin", color="D0D0D0"),
    right=Side(style="thin", color="D0D0D0"),
    top=Side(style="thin", color="D0D0D0"),
    bottom=Side(style="thin", color="D0D0D0"),
)

ALIGN_CENTER = Alignment(horizontal="center", vertical="center")
ALIGN_LEFT = Alignment(horizontal="left", vertical="center")
ALIGN_RIGHT = Alignment(horizontal="right", vertical="center")

FMT_EURO = '#,##0.00 €'
FMT_HOURS = '0.0'

# ---------------------------------------------------------------------------
# Date della settimana corrente (Lun-Dom)
# ---------------------------------------------------------------------------
today = date.today()
monday = today - timedelta(days=today.weekday())
week_dates = [monday + timedelta(days=i) for i in range(7)]
DAY_NAMES = ["Lun", "Mar", "Mer", "Gio", "Ven", "Sab", "Dom"]

# ---------------------------------------------------------------------------
# Dati di esempio
# ---------------------------------------------------------------------------
SAMPLE_DATA = [
    ("Acme Srl", "Progetto Web 2026", "Sviluppo", 50, [8, 7.5, 8, 6, 4, 0, 0]),
    ("Acme Srl", "Manutenzione sito", "Manutenzione", 40, [2, 0, 1, 0, 2, 0, 0]),
    ("Beta Corp", "Consulenza CRM", "Consulenza", 65, [0, 4, 0, 8, 4, 0, 0]),
    ("Beta Corp", "Formazione team", "Formazione", 55, [0, 0, 4, 0, 0, 0, 0]),
    ("Gamma Design", "Supporto tecnico", "Supporto", 35, [1, 1, 1, 1, 1, 0, 0]),
]

TYPES = ["Sviluppo", "Manutenzione", "Consulenza", "Formazione", "Supporto", "Altro"]
CLIENTS = ["Acme Srl", "Beta Corp", "Gamma Design"]

DATA_START_ROW = 5
DATA_END_ROW = 24
TOTAL_ROW = 25


def build_timesheet(wb: Workbook):
    ws = wb.active
    ws.title = "Timesheet"
    ws.sheet_properties.tabColor = COLOR_HEADER

    # -- Larghezze colonne ------------------------------------------------
    col_widths = {
        "A": 18, "B": 22, "C": 14, "D": 9,
        "E": 9, "F": 9, "G": 9, "H": 9, "I": 9, "J": 9, "K": 9,
        "L": 11, "M": 12,
        "N": 16, "O": 16, "P": 16, "Q": 16, "R": 16, "S": 16, "T": 16,
    }
    for col, w in col_widths.items():
        ws.column_dimensions[col].width = w

    # ── HEADER ────────────────────────────────────────────────────────────
    # Riga 1: titolo
    ws.merge_cells("A1:T1")
    c = ws["A1"]
    c.value = "TIMESHEET SETTIMANALE"
    c.font = FONT_TITLE
    c.fill = FILL_HEADER
    c.alignment = ALIGN_CENTER
    ws.row_dimensions[1].height = 38
    for col_idx in range(2, 21):
        cell = ws.cell(row=1, column=col_idx)
        cell.fill = FILL_HEADER

    # Riga 2: sottotitolo settimana
    ws.merge_cells("A2:T2")
    ws["A2"].value = (
        f"Settimana: Lun {week_dates[0].strftime('%d/%m/%Y')} "
        f"– Dom {week_dates[6].strftime('%d/%m/%Y')}"
    )
    ws["A2"].font = FONT_SUBTITLE
    ws["A2"].alignment = ALIGN_CENTER
    ws.row_dimensions[2].height = 22

    # Riga 3: separatore
    ws.row_dimensions[3].height = 8

    # ── INTESTAZIONI COLONNE (riga 4) ─────────────────────────────────────
    headers = ["Cliente", "Commessa", "Tipo", "€/h"]
    for i, d in enumerate(week_dates):
        headers.append(f"{DAY_NAMES[i]}\n{d.strftime('%d/%m')}")
    headers += ["Tot Ore", "Tot €"]
    for i, d in enumerate(week_dates):
        headers.append(f"Note {DAY_NAMES[i]}")

    for col_idx, h in enumerate(headers, 1):
        cell = ws.cell(row=4, column=col_idx)
        cell.value = h
        cell.font = FONT_HEADER
        cell.fill = FILL_HEADER
        cell.alignment = Alignment(
            horizontal="center", vertical="center", wrap_text=True
        )
        cell.border = THIN_BORDER
    ws.row_dimensions[4].height = 32

    # ── AREA DATI (righe 5-24) ────────────────────────────────────────────
    for row in range(DATA_START_ROW, DATA_END_ROW + 1):
        idx = row - DATA_START_ROW
        is_alt = idx % 2 == 1
        bg = FILL_ALT if is_alt else FILL_WHITE

        for col_idx in range(1, 21):
            cell = ws.cell(row=row, column=col_idx)
            cell.font = FONT_DATA
            cell.border = THIN_BORDER
            # sfondo weekend per colonne J (10) e K (11)
            if col_idx in (10, 11):
                cell.fill = FILL_WEEKEND
            else:
                cell.fill = bg

            # Allineamento
            if col_idx <= 2:
                cell.alignment = ALIGN_LEFT
            elif col_idx == 3:
                cell.alignment = ALIGN_CENTER
            else:
                cell.alignment = ALIGN_CENTER

        # Formato colonna D (€/h)
        ws.cell(row=row, column=4).number_format = FMT_EURO
        # Formato ore (E-K)
        for c in range(5, 12):
            ws.cell(row=row, column=c).number_format = FMT_HOURS
        # Formato Tot Ore (L)
        ws.cell(row=row, column=12).number_format = FMT_HOURS
        # Formato Tot € (M)
        ws.cell(row=row, column=13).number_format = FMT_EURO

        # Formule Tot Ore e Tot €
        ws.cell(row=row, column=12).value = f"=SUM(E{row}:K{row})"
        ws.cell(row=row, column=13).value = f"=L{row}*D{row}"

    # Dati di esempio
    for i, (cliente, commessa, tipo, tariffa, ore) in enumerate(SAMPLE_DATA):
        row = DATA_START_ROW + i
        ws.cell(row=row, column=1).value = cliente
        ws.cell(row=row, column=2).value = commessa
        ws.cell(row=row, column=3).value = tipo
        ws.cell(row=row, column=4).value = tariffa
        for j, h in enumerate(ore):
            ws.cell(row=row, column=5 + j).value = h

    # ── RIGA TOTALE (25) ──────────────────────────────────────────────────
    ws.row_dimensions[TOTAL_ROW].height = 28
    for col_idx in range(1, 21):
        cell = ws.cell(row=TOTAL_ROW, column=col_idx)
        cell.fill = FILL_GREEN
        cell.border = THIN_BORDER
        cell.font = FONT_TOTAL_VALUE
        cell.alignment = ALIGN_CENTER
    ws.cell(row=TOTAL_ROW, column=1).value = "TOTALE"
    ws.cell(row=TOTAL_ROW, column=1).font = FONT_TOTAL_LABEL
    ws.cell(row=TOTAL_ROW, column=1).alignment = ALIGN_LEFT
    ws.merge_cells(f"A{TOTAL_ROW}:C{TOTAL_ROW}")

    # Totali per ogni colonna giorno
    for col_idx in range(5, 12):
        col_l = get_column_letter(col_idx)
        ws.cell(row=TOTAL_ROW, column=col_idx).value = (
            f"=SUM({col_l}{DATA_START_ROW}:{col_l}{DATA_END_ROW})"
        )
        ws.cell(row=TOTAL_ROW, column=col_idx).number_format = FMT_HOURS

    ws.cell(row=TOTAL_ROW, column=12).value = (
        f"=SUM(L{DATA_START_ROW}:L{DATA_END_ROW})"
    )
    ws.cell(row=TOTAL_ROW, column=12).number_format = FMT_HOURS
    ws.cell(row=TOTAL_ROW, column=13).value = (
        f"=SUM(M{DATA_START_ROW}:M{DATA_END_ROW})"
    )
    ws.cell(row=TOTAL_ROW, column=13).number_format = FMT_EURO

    # ── DATA VALIDATION (dropdown Tipo) ───────────────────────────────────
    dv = DataValidation(
        type="list",
        formula1='"' + ",".join(TYPES) + '"',
        allow_blank=True,
    )
    dv.error = "Scegli un tipo valido"
    dv.errorTitle = "Tipo non valido"
    dv.prompt = "Seleziona il tipo di attività"
    dv.promptTitle = "Tipo Attività"
    ws.add_data_validation(dv)
    dv.add(f"C{DATA_START_ROW}:C{DATA_END_ROW}")

    # ── FREEZE PANES ──────────────────────────────────────────────────────
    ws.freeze_panes = "A5"

    return ws


def build_dashboard(wb: Workbook):
    ws = wb.create_sheet("Dashboard")
    ws.sheet_properties.tabColor = COLOR_GREEN

    col_widths = {
        "A": 20, "B": 14, "C": 14, "D": 12, "E": 4,
        "F": 20, "G": 14, "H": 14, "I": 12,
    }
    for col, w in col_widths.items():
        ws.column_dimensions[col].width = w

    # ── TITOLO ────────────────────────────────────────────────────────────
    ws.merge_cells("A1:I1")
    ws["A1"].value = "DASHBOARD SETTIMANALE"
    ws["A1"].font = Font(name="Calibri", size=16, bold=True, color=COLOR_HEADER)
    ws["A1"].alignment = ALIGN_CENTER
    ws.row_dimensions[1].height = 34
    for col_idx in range(1, 10):
        ws.cell(row=1, column=col_idx).border = Border(
            bottom=Side(style="medium", color=COLOR_HEADER)
        )

    # ── KPI BOXES (riga 3-4) ──────────────────────────────────────────────
    ts = "Timesheet"
    kpi_defs = [
        ("Totale Ore", f"={ts}!L{TOTAL_ROW}", FMT_HOURS, COLOR_GREEN, "A", "B"),
        ("Totale Fatturato", f"={ts}!M{TOTAL_ROW}", FMT_EURO, COLOR_ORANGE, "C", "D"),
        (
            "Ore Medie / Giorno",
            f"=IF({ts}!L{TOTAL_ROW}=0,0,{ts}!L{TOTAL_ROW}/7)",
            FMT_HOURS,
            COLOR_HEADER,
            "F",
            "G",
        ),
        (
            "Clienti Attivi",
            f'=SUMPRODUCT((LEN(TRIM({ts}!A{DATA_START_ROW}:A{DATA_END_ROW}))>0)'
            f'*(1/COUNTIF({ts}!A{DATA_START_ROW}:A{DATA_END_ROW},'
            f'{ts}!A{DATA_START_ROW}:A{DATA_END_ROW})))',
            "0",
            "8E44AD",
            "H",
            "I",
        ),
    ]
    for title, formula, fmt, color, col_start, col_end in kpi_defs:
        ws.merge_cells(f"{col_start}3:{col_end}3")
        cell_title = ws[f"{col_start}3"]
        cell_title.value = title
        cell_title.font = FONT_KPI_TITLE
        cell_title.alignment = ALIGN_CENTER
        fill_light = PatternFill("solid", fgColor="F7F9FC")
        cell_title.fill = fill_light
        ws[f"{col_end}3"].fill = fill_light

        ws.merge_cells(f"{col_start}4:{col_end}4")
        cell_val = ws[f"{col_start}4"]
        cell_val.value = formula
        cell_val.font = Font(name="Calibri", size=18, bold=True, color=color)
        cell_val.number_format = fmt
        cell_val.alignment = ALIGN_CENTER
        cell_val.fill = fill_light
        ws[f"{col_end}4"].fill = fill_light

        # Bordi KPI box
        for r in (3, 4):
            for c_letter in (col_start, col_end):
                ws[f"{c_letter}{r}"].border = Border(
                    top=Side(style="thin", color=color),
                    bottom=Side(style="thin", color=color),
                    left=Side(style="thin", color=color),
                    right=Side(style="thin", color=color),
                )
    ws.row_dimensions[3].height = 22
    ws.row_dimensions[4].height = 36

    # ── RIEPILOGO PER CLIENTE (righe 6-) ──────────────────────────────────
    r = 6
    ws.cell(row=r, column=1).value = "Riepilogo per Cliente"
    ws.cell(row=r, column=1).font = FONT_DASH_SECTION
    r += 1

    client_headers = ["Cliente", "Ore Totali", "Fatturato €", "% Ore"]
    for ci, h in enumerate(client_headers, 1):
        cell = ws.cell(row=r, column=ci)
        cell.value = h
        cell.font = FONT_DASH_HEADER
        cell.fill = FILL_HEADER
        cell.alignment = ALIGN_CENTER
        cell.border = THIN_BORDER
    r += 1
    client_start_row = r

    for cli in CLIENTS:
        ws.cell(row=r, column=1).value = cli
        ws.cell(row=r, column=1).font = FONT_DATA
        ws.cell(row=r, column=1).border = THIN_BORDER
        ws.cell(row=r, column=1).alignment = ALIGN_LEFT

        # Ore Totali = SUMIF
        ws.cell(row=r, column=2).value = (
            f'=SUMPRODUCT(({ts}!A{DATA_START_ROW}:A{DATA_END_ROW}=A{r})'
            f'*{ts}!L{DATA_START_ROW}:L{DATA_END_ROW})'
        )
        ws.cell(row=r, column=2).number_format = FMT_HOURS
        ws.cell(row=r, column=2).font = FONT_DATA
        ws.cell(row=r, column=2).border = THIN_BORDER
        ws.cell(row=r, column=2).alignment = ALIGN_CENTER

        # Fatturato €
        ws.cell(row=r, column=3).value = (
            f'=SUMPRODUCT(({ts}!A{DATA_START_ROW}:A{DATA_END_ROW}=A{r})'
            f'*{ts}!M{DATA_START_ROW}:M{DATA_END_ROW})'
        )
        ws.cell(row=r, column=3).number_format = FMT_EURO
        ws.cell(row=r, column=3).font = FONT_DATA
        ws.cell(row=r, column=3).border = THIN_BORDER
        ws.cell(row=r, column=3).alignment = ALIGN_CENTER

        # % Ore
        ws.cell(row=r, column=4).value = (
            f'=IF({ts}!L{TOTAL_ROW}=0,0,B{r}/{ts}!L{TOTAL_ROW})'
        )
        ws.cell(row=r, column=4).number_format = "0.0%"
        ws.cell(row=r, column=4).font = FONT_DATA
        ws.cell(row=r, column=4).border = THIN_BORDER
        ws.cell(row=r, column=4).alignment = ALIGN_CENTER

        if (r - client_start_row) % 2 == 1:
            for ci in range(1, 5):
                ws.cell(row=r, column=ci).fill = FILL_ALT
        r += 1
    client_end_row = r - 1

    # ── RIEPILOGO PER TIPO (colonne F-I, stesse righe) ────────────────────
    tr = 6
    ws.cell(row=tr, column=6).value = "Riepilogo per Tipo Attività"
    ws.cell(row=tr, column=6).font = FONT_DASH_SECTION
    tr += 1

    type_headers = ["Tipo", "Ore", "Fatturato €", "% Ore"]
    for ci, h in enumerate(type_headers):
        cell = ws.cell(row=tr, column=6 + ci)
        cell.value = h
        cell.font = FONT_DASH_HEADER
        cell.fill = FILL_HEADER
        cell.alignment = ALIGN_CENTER
        cell.border = THIN_BORDER
    tr += 1
    type_start_row = tr

    for tipo in TYPES:
        ws.cell(row=tr, column=6).value = tipo
        ws.cell(row=tr, column=6).font = FONT_DATA
        ws.cell(row=tr, column=6).border = THIN_BORDER
        ws.cell(row=tr, column=6).alignment = ALIGN_LEFT

        ws.cell(row=tr, column=7).value = (
            f'=SUMPRODUCT(({ts}!C{DATA_START_ROW}:C{DATA_END_ROW}=F{tr})'
            f'*{ts}!L{DATA_START_ROW}:L{DATA_END_ROW})'
        )
        ws.cell(row=tr, column=7).number_format = FMT_HOURS
        ws.cell(row=tr, column=7).font = FONT_DATA
        ws.cell(row=tr, column=7).border = THIN_BORDER
        ws.cell(row=tr, column=7).alignment = ALIGN_CENTER

        ws.cell(row=tr, column=8).value = (
            f'=SUMPRODUCT(({ts}!C{DATA_START_ROW}:C{DATA_END_ROW}=F{tr})'
            f'*{ts}!M{DATA_START_ROW}:M{DATA_END_ROW})'
        )
        ws.cell(row=tr, column=8).number_format = FMT_EURO
        ws.cell(row=tr, column=8).font = FONT_DATA
        ws.cell(row=tr, column=8).border = THIN_BORDER
        ws.cell(row=tr, column=8).alignment = ALIGN_CENTER

        ws.cell(row=tr, column=9).value = (
            f'=IF({ts}!L{TOTAL_ROW}=0,0,G{tr}/{ts}!L{TOTAL_ROW})'
        )
        ws.cell(row=tr, column=9).number_format = "0.0%"
        ws.cell(row=tr, column=9).font = FONT_DATA
        ws.cell(row=tr, column=9).border = THIN_BORDER
        ws.cell(row=tr, column=9).alignment = ALIGN_CENTER

        if (tr - type_start_row) % 2 == 1:
            for ci in range(6, 10):
                ws.cell(row=tr, column=ci).fill = FILL_ALT
        tr += 1
    type_end_row = tr - 1

    # ── DISTRIBUZIONE GIORNALIERA (dati per grafico) ──────────────────────
    daily_row = max(client_end_row, type_end_row) + 3
    ws.cell(row=daily_row, column=1).value = "Distribuzione Giornaliera"
    ws.cell(row=daily_row, column=1).font = FONT_DASH_SECTION
    daily_row += 1

    day_cols_ts = ["E", "F", "G", "H", "I", "J", "K"]  # colonne giorni in Timesheet
    ws.cell(row=daily_row, column=1).value = "Giorno"
    ws.cell(row=daily_row, column=1).font = FONT_DASH_HEADER
    ws.cell(row=daily_row, column=1).fill = FILL_HEADER
    ws.cell(row=daily_row, column=1).alignment = ALIGN_CENTER
    ws.cell(row=daily_row, column=1).border = THIN_BORDER
    ws.cell(row=daily_row, column=2).value = "Ore"
    ws.cell(row=daily_row, column=2).font = FONT_DASH_HEADER
    ws.cell(row=daily_row, column=2).fill = FILL_HEADER
    ws.cell(row=daily_row, column=2).alignment = ALIGN_CENTER
    ws.cell(row=daily_row, column=2).border = THIN_BORDER
    daily_row += 1
    daily_start_row = daily_row

    for i, dn in enumerate(DAY_NAMES):
        ws.cell(row=daily_row, column=1).value = dn
        ws.cell(row=daily_row, column=1).font = FONT_DATA
        ws.cell(row=daily_row, column=1).border = THIN_BORDER
        ws.cell(row=daily_row, column=1).alignment = ALIGN_CENTER
        col_l = day_cols_ts[i]
        ws.cell(row=daily_row, column=2).value = (
            f"=SUM({ts}!{col_l}{DATA_START_ROW}:{col_l}{DATA_END_ROW})"
        )
        ws.cell(row=daily_row, column=2).number_format = FMT_HOURS
        ws.cell(row=daily_row, column=2).font = FONT_DATA
        ws.cell(row=daily_row, column=2).border = THIN_BORDER
        ws.cell(row=daily_row, column=2).alignment = ALIGN_CENTER
        if i >= 5:
            ws.cell(row=daily_row, column=1).fill = FILL_WEEKEND
            ws.cell(row=daily_row, column=2).fill = FILL_WEEKEND
        daily_row += 1
    daily_end_row = daily_row - 1

    # ── GRAFICI ───────────────────────────────────────────────────────────
    chart_top_row = daily_end_row + 3

    # 1. Ore per Cliente — BarChart orizzontale
    c1 = BarChart()
    c1.type = "bar"
    c1.style = 10
    c1.title = "Ore per Cliente"
    c1.y_axis.title = None
    c1.x_axis.title = "Ore"
    data1 = Reference(ws, min_col=2, min_row=client_start_row - 1,
                      max_row=client_end_row)
    cats1 = Reference(ws, min_col=1, min_row=client_start_row,
                      max_row=client_end_row)
    c1.add_data(data1, titles_from_data=True)
    c1.set_categories(cats1)
    c1.shape = 4
    c1.width = 18
    c1.height = 11
    s1 = c1.series[0]
    s1.graphicalProperties.solidFill = COLOR_GREEN
    ws.add_chart(c1, f"A{chart_top_row}")

    # 2. Fatturato per Cliente — BarChart orizzontale
    c2 = BarChart()
    c2.type = "bar"
    c2.style = 10
    c2.title = "Fatturato per Cliente"
    c2.y_axis.title = None
    c2.x_axis.title = "€"
    data2 = Reference(ws, min_col=3, min_row=client_start_row - 1,
                      max_row=client_end_row)
    cats2 = Reference(ws, min_col=1, min_row=client_start_row,
                      max_row=client_end_row)
    c2.add_data(data2, titles_from_data=True)
    c2.set_categories(cats2)
    c2.width = 18
    c2.height = 11
    s2 = c2.series[0]
    s2.graphicalProperties.solidFill = COLOR_ORANGE
    ws.add_chart(c2, f"F{chart_top_row}")

    chart_mid_row = chart_top_row + 16

    # 3. Ore per Tipo Attività — PieChart
    c3 = PieChart()
    c3.title = "Ore per Tipo Attività"
    c3.style = 10
    data3 = Reference(ws, min_col=7, min_row=type_start_row - 1,
                      max_row=type_end_row)
    cats3 = Reference(ws, min_col=6, min_row=type_start_row,
                      max_row=type_end_row)
    c3.add_data(data3, titles_from_data=True)
    c3.set_categories(cats3)
    c3.width = 18
    c3.height = 13
    c3.dataLabels = DataLabelList()
    c3.dataLabels.showPercent = True
    c3.dataLabels.showCatName = True
    c3.dataLabels.showVal = False
    # Colori fette
    pie_colors = ["27AE60", "E67E22", "2980B9", "8E44AD", "E74C3C", "95A5A6"]
    for idx, color in enumerate(pie_colors):
        pt = DataPoint(idx=idx)
        pt.graphicalProperties.solidFill = color
        c3.series[0].data_points.append(pt)
    ws.add_chart(c3, f"A{chart_mid_row}")

    # 4. Distribuzione Settimanale — BarChart verticale
    c4 = BarChart()
    c4.type = "col"
    c4.style = 10
    c4.title = "Distribuzione Settimanale"
    c4.y_axis.title = "Ore"
    c4.x_axis.title = None
    data4 = Reference(ws, min_col=2, min_row=daily_start_row - 1,
                       max_row=daily_end_row)
    cats4 = Reference(ws, min_col=1, min_row=daily_start_row,
                       max_row=daily_end_row)
    c4.add_data(data4, titles_from_data=True)
    c4.set_categories(cats4)
    c4.width = 18
    c4.height = 13
    s4 = c4.series[0]
    s4.graphicalProperties.solidFill = COLOR_HEADER
    ws.add_chart(c4, f"F{chart_mid_row}")

    return ws


def main():
    wb = Workbook()
    build_timesheet(wb)
    build_dashboard(wb)

    output = "timesheet.xlsx"
    wb.save(output)
    print(f"✓ File generato: {output}")
    print(f"  Settimana: {week_dates[0].strftime('%d/%m/%Y')} – {week_dates[6].strftime('%d/%m/%Y')}")
    print(f"  Fogli: Timesheet, Dashboard")
    print(f"  Righe dati: {DATA_START_ROW}-{DATA_END_ROW} ({DATA_END_ROW - DATA_START_ROW + 1} righe)")
    print(f"  Grafici dashboard: 4")


if __name__ == "__main__":
    main()
